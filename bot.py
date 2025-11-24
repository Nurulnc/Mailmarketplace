import telebot
from telebot import types
import os
import openpyxl
from io import BytesIO

# === তোমার তথ্য ===
TOKEN = "8594094725:AAEtkG2hAgpn7oNxtp8uvrBiFwcaZ2d-oKA"
ADMIN_ID = 1651695602
PRICE_PER_MAIL = 2

PAYMENT_INFO = """💳 Payment Methods:
🔴 bKash: 01815243007
🟢 Binance Pay: 38017799
**Total Amount: {total} Taka** ({quantity} × {price} Tk per mail)
📤 Send **screenshot** after payment."""

user_data = {}  # {user_id: {'quantity': 5, 'total': 15, 'state': '...', 'admin_msg_id': ...}}
pending_approvals = {}  # {user_id: {'quantity': X, 'chat_id': Y}}  -> যখন admin ফাইল পাঠাবে তখন এটা থেকে নেব

bot = telebot.TeleBot(TOKEN)

@bot.message_handler(commands=['start'])
def start(message):
    markup = types.InlineKeyboardMarkup()
    btn_order = types.InlineKeyboardButton("🛒 Buy .EDU Email", callback_data="order")
    markup.add(btn_order)
   
    bot.send_message(message.chat.id,
                     "🌟 **.EDU Email Seller Bot** 🌟\n\n"
                     "💰 **Price: 2 Taka per mail | 24hr live**\n"
                     "✅ Instant delivery after payment\n"
                     "🚀 GitHub, Spotify, Office 365, etc.",
                     parse_mode="Markdown", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "order")
def order_callback(call):
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id,
                          text="📋 Order .EDU Email")
   
    user_data[call.from_user.id] = {"state": "waiting_quantity"}
    bot.send_message(call.message.chat.id,
                     "📦 **কতগুলো .EDU মেইল কিনবেন?**\n\n"
                     "শুধু সংখ্যা লিখুন (যেমন: `5`)",
                     parse_mode="Markdown")

@bot.message_handler(func=lambda m: m.from_user.id in user_data and user_data[m.from_user.id]["state"] == "waiting_quantity")
def handle_quantity(message):
    user_id = message.from_user.id
    try:
        quantity = int(message.text.strip())
        if quantity < 1 or quantity > 500:  # লিমিট দিলাম
            raise ValueError
    except:
        bot.send_message(user_id, "❌ শুধু ১-৫০০ এর মধ্যে সংখ্যা লিখুন।")
        return
   
    total = quantity * PRICE_PER_MAIL
    user_data[user_id].update({
        "quantity": quantity,
        "total": total,
        "state": "waiting_screenshot"
    })
   
    bot.send_message(user_id, PAYMENT_INFO.format(total=total, quantity=quantity, price=PRICE_PER_MAIL),
                     parse_mode="Markdown")
    bot.send_message(user_id, "📤 এখন **পেমেন্ট স্ক্রিনশট** পাঠান।", parse_mode="Markdown")

@bot.message_handler(content_types=['photo'],
                     func=lambda m: m.from_user.id in user_data and user_data[m.from_user.id]["state"] == "waiting_screenshot")
def handle_photo(message):
    user_id = message.from_user.id
    data = user_data[user_id]
    quantity = data["quantity"]
    total = data["total"]
   
    forwarded = bot.forward_message(ADMIN_ID, message.chat.id, message.message_id)
   
    username = message.from_user.username or "No username"
    full_name = f"{message.from_user.first_name} {message.from_user.last_name or ''}".strip()
   
    admin_text = (f"🟢 **NEW ORDER** 🟢\n\n"
                  f"👤 **User**: {full_name}\n"
                  f"🆔 **ID**: <code>{user_id}</code>\n"
                  f"✏️ **Username**: @{username}\n"
                  f"📦 **Quantity**: {quantity} mail(s)\n"
                  f"💰 **Total**: {total} Taka\n\n"
                  f"📸 Screenshot পাওয়া গেছে। এখন Approve করতে নিচের কমান্ড ব্যবহার করো:\n\n"
                  f"<code>/approve {user_id} {quantity}</code>\n"
                  f"এরপর .txt বা .xlsx ফাইল পাঠাও।")
   
    sent = bot.send_message(ADMIN_ID, admin_text, parse_mode="HTML", reply_to_message_id=forwarded.message_id)
   
    bot.send_message(user_id, "✅ স্ক্রিনশট পাওয়া গেছে!\n\n⏳ এডমিন পেমেন্ট চেক করছেন...")
   
    # Save for approval
    pending_approvals[user_id] = {
        "quantity": quantity,
        "chat_id": message.chat.id,
        "admin_notify_msg_id": sent.message_id
    }
    user_data.pop(user_id, None)

# ====== ADMIN APPROVE SYSTEM (File Based) ======
@bot.message_handler(commands=['approve'])
def approve_command(message):
    if message.from_user.id != ADMIN_ID:
        return
    try:
        parts = message.text.split()
        if len(parts) != 3:
            raise ValueError
        user_id = int(parts[1])
        qty = int(parts[2])
        
        if user_id not in pending_approvals or pending_approvals[user_id]["quantity"] != qty:
            bot.reply_to(message, "❌ এই ইউজারের অর্ডার ম্যাচ করছে না। আবার চেক করো।")
            return
            
        bot.reply_to(message, f"✅ ঠিক আছে! এখন {qty}টা মেইলের একটা .txt অথবা .xlsx ফাইল পাঠাও।\n"
                              "প্রতি লাইনে একটা → email:password")
        # Store temporary state
        pending_approvals[user_id]["waiting_file"] = True
        pending_approvals[user_id]["approve_msg_id"] = message.message_id
        
    except:
        bot.reply_to(message, "❌ ভুল ফরম্যাট!\n\nসঠিক: <code>/approve user_id quantity</code>", parse_mode="HTML")

# Handle TXT or XLSX file from Admin
@bot.message_handler(content_types=['document'])
def handle_admin_document(message):
    if message.from_user.id != ADMIN_ID:
        return
        
    if not message.document.file_name.lower().endswith(('.txt', '.xlsx')):
        return
        
    # Check if any pending approval waiting for file
    target_user_id = None
    for uid, data in pending_approvals.items():
        if data.get("waiting_file"):
            target_user_id = uid
            break
    if not target_user_id:
        bot.reply_to(message, "❌ কোনো অর্ডারের জন্য ফাইল অপেক্ষা করছে না। প্রথমে /approve দাও।")
        return

    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    required_qty = pending_approvals[target_user_id]["quantity"]
    mails = []

    if message.document.file_name.lower().endswith('.txt'):
        text = downloaded_file.decode('utf-8')
        mails = [line.strip() for line in text.splitlines() if ':' in line]
    
    elif message.document.file_name.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(BytesIO(downloaded_file))
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] and ':' in str(row[0]):
                mails.append(str(row[0]).strip())

    if len(mails) != required_qty:
        bot.reply_to(message, f"❌ ভুল পরিমাণ! চেয়েছে {required_qty}টা, পাওয়া গেছে {len(mails)}টা।")
        return

    # Success — Send to buyer
    chat_id = pending_approvals[target_user_id]["chat_id"]
    
    # Send as TXT file
    txt_content = "\n".join(mails)
    bio = BytesIO(txt_content.encode('utf-8'))
    bio.name = f"EDU_Emails_{target_user_id}.txt"
    bot.send_document(chat_id, bio, caption="🎉 **তোমার .EDU মেইলগুলো এসেছে!**\n\n"
                                              "🔐 তৎক্ষণাৎ পাসওয়ার্ড চেঞ্জ করো!\n"
                                              "❤️ ধন্যবাদ কেনার জন্য!", parse_mode="Markdown")
    
    # Also send as Excel (optional)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EDU Emails"
    ws.append(["Email:Password"])
    for mail in mails:
        ws.append([mail])
    bio2 = BytesIO()
    wb.save(bio2)
    bio2.name = f"EDU_Emails_{target_user_id}.xlsx"
    bio2.seek(0)
    bot.send_document(chat_id, bio2, caption="📊 এক্সেল ফাইলও দিলাম (সুবিধার জন্য)")

    # Notify Admin
    bot.reply_to(message, f"✅ সফল! {required_qty}টা মেইল পাঠানো হয়েছে → {target_user_id}")
    
    # Clean up
    del pending_approvals[target_user_id]

# Fallback
@bot.message_handler(func=lambda m: True)
def fallback(message):
    if message.from_user.id not in user_data and message.from_user.id not in pending_approvals:
        bot.send_message(message.chat.id, "👋 /start দিয়ে শুরু করো ভাই।")

print("Bot চালু হয়েছে...")
bot.infinity_polling()
